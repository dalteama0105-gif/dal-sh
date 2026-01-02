import tkinter as tk
from tkinter import ttk, messagebox, filedialog, Menu
import sqlite3
from datetime import datetime, timedelta
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

class AttendanceApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Attendance Management System")
        self.root.geometry("1100x750")
        self.root.minsize(900, 600)
        
        # Session state
        self.active_session = None
        self.session_timer = None
        self.session_end_time = None
        self.late_end_time = None
        self.current_phase = None 
        
        self.scan_map = {}
        # Sorting state: {column_name: bool_is_ascending}
        self.sort_state = {"Name": True, "DOB": True, "Age": True, "Role": True}
        
        self.init_database()
        self.create_ui()
        
    def init_database(self):
        """Initialize SQLite database with required tables"""
        self.conn = sqlite3.connect('AttendanceManagementSystem.db')
        self.cursor = self.conn.cursor()
        self.cursor.execute("PRAGMA journal_mode=WAL")
        self.cursor.execute("PRAGMA foreign_keys = ON")
        
        self.cursor.execute('''
            CREATE TABLE IF NOT EXISTS people (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                key TEXT UNIQUE NOT NULL,
                name TEXT NOT NULL,
                dob TEXT NOT NULL,
                age INTEGER,
                personality TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        
        self.cursor.execute('''
            CREATE TABLE IF NOT EXISTS sessions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                starter_key TEXT NOT NULL,
                title TEXT NOT NULL,
                description TEXT,
                date_str TEXT,
                start_time TIMESTAMP NOT NULL,
                normal_end_time TIMESTAMP,
                final_end_time TIMESTAMP,
                normal_duration_sec INTEGER,
                late_duration_sec INTEGER,
                status TEXT DEFAULT 'active',
                FOREIGN KEY (starter_key) REFERENCES people(key) ON DELETE CASCADE
            )
        ''')
        
        self.cursor.execute('''
            CREATE TABLE IF NOT EXISTS attendance (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                session_id INTEGER NOT NULL,
                person_key TEXT NOT NULL,
                scan_time TIMESTAMP NOT NULL,
                status TEXT NOT NULL, 
                FOREIGN KEY (session_id) REFERENCES sessions(id) ON DELETE CASCADE,
                FOREIGN KEY (person_key) REFERENCES people(key) ON DELETE CASCADE
            )
        ''')
        self.conn.commit()
        
    def create_ui(self):
        style = ttk.Style()
        style.theme_use('clam')
        style.map('Treeview', background=[('selected', '#2196F3')])
        
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill='both', expand=True, padx=10, pady=10)
        
        self.scan_frame = ttk.Frame(self.notebook)
        self.new_frame = ttk.Frame(self.notebook)
        self.history_frame = ttk.Frame(self.notebook)
        self.list_frame = ttk.Frame(self.notebook)
        
        self.notebook.add(self.scan_frame, text='🔄 Record Attendance')
        self.notebook.add(self.history_frame, text='🗂️ Past Records')
        self.notebook.add(self.new_frame, text='➕ New (Add Person)')
        self.notebook.add(self.list_frame, text='👥 List (View People)')
        
        self.build_scan_tab()
        self.build_history_tab()
        self.build_new_tab()
        self.build_list_tab()
        
    def build_scan_tab(self):
        main_container = ttk.Frame(self.scan_frame)
        main_container.pack(fill='both', expand=True, padx=20, pady=10)
        
        tk.Label(main_container, text="Live Attendance Session", font=('Arial', 16, 'bold')).pack(pady=5)
        
        self.setup_frame = ttk.LabelFrame(main_container, text="Start New Session", padding=15)
        self.setup_frame.pack(fill='x', pady=10)
        
        row1 = ttk.Frame(self.setup_frame); row1.pack(fill='x', pady=5)
        ttk.Label(row1, text="Starter Person Key:", font=('Arial', 10, 'bold'), width=18).pack(side='left')
        self.starter_key_entry = ttk.Entry(row1, font=('Arial', 10), width=30); self.starter_key_entry.pack(side='left', padx=5)
        
        row2 = ttk.Frame(self.setup_frame); row2.pack(fill='x', pady=5)
        ttk.Label(row2, text="Session Title:", width=18).pack(side='left')
        self.session_title_entry = ttk.Entry(row2, font=('Arial', 10), width=30); self.session_title_entry.pack(side='left', padx=5)
        ttk.Label(row2, text="Desc:", width=6).pack(side='left', padx=5)
        self.session_desc_entry = ttk.Entry(row2, font=('Arial', 10), width=30); self.session_desc_entry.pack(side='left', padx=5)
        
        time_frame = ttk.LabelFrame(self.setup_frame, text="Time Configuration", padding=10); time_frame.pack(fill='x', pady=10)
        t_choice_row = ttk.Frame(time_frame); t_choice_row.pack(fill='x', pady=5)
        ttk.Label(t_choice_row, text="⏱️ Main Attendance:", width=18, font=('Arial', 10, 'bold')).pack(side='left')
        self.time_mode = tk.StringVar(value="countdown")
        ttk.Radiobutton(t_choice_row, text="Countdown", variable=self.time_mode, value="countdown").pack(side='left', padx=10)
        ttk.Radiobutton(t_choice_row, text="End Time", variable=self.time_mode, value="clock").pack(side='left', padx=10)

        t_input_row = ttk.Frame(time_frame); t_input_row.pack(fill='x', pady=5)
        self.countdown_frame = ttk.Frame(t_input_row); self.countdown_frame.pack(side='left', padx=20)
        self.main_min = tk.StringVar(value="30"); self.main_sec = tk.StringVar(value="0")
        ttk.Spinbox(self.countdown_frame, from_=0, to=180, textvariable=self.main_min, width=5).pack(side='left')
        ttk.Label(self.countdown_frame, text="min").pack(side='left', padx=2)
        ttk.Spinbox(self.countdown_frame, from_=0, to=59, textvariable=self.main_sec, width=5).pack(side='left')
        ttk.Label(self.countdown_frame, text="sec").pack(side='left', padx=2)
        
        self.clock_frame = ttk.Frame(t_input_row); self.clock_frame.pack(side='left', padx=20)
        ttk.Label(self.clock_frame, text="End Time (HH:MM):").pack(side='left', padx=5)
        self.end_hour = ttk.Spinbox(self.clock_frame, from_=0, to=23, width=5, format="%02.0f"); self.end_hour.set(datetime.now().hour); self.end_hour.pack(side='left', padx=2)
        ttk.Label(self.clock_frame, text=":").pack(side='left')
        self.end_minute = ttk.Spinbox(self.clock_frame, from_=0, to=59, width=5, format="%02.0f"); self.end_minute.set(datetime.now().minute); self.end_minute.pack(side='left', padx=2)
        
        ttk.Label(self.clock_frame, text="  |  Current:", foreground="gray").pack(side='left', padx=5)
        self.rt_clock_label = ttk.Label(self.clock_frame, text="00:00:00", font=('Arial', 10, 'bold'), foreground="#2196F3"); self.rt_clock_label.pack(side='left')
        self.tick_clock() 
        
        l_row = ttk.Frame(time_frame); l_row.pack(fill='x', pady=5)
        ttk.Label(l_row, text="⏰ Late Clock-In:", width=18, font=('Arial', 10, 'bold')).pack(side='left')
        self.late_min = tk.StringVar(value="10"); self.late_sec = tk.StringVar(value="0")
        ttk.Spinbox(l_row, from_=0, to=60, textvariable=self.late_min, width=5).pack(side='left')
        ttk.Label(l_row, text="min").pack(side='left', padx=2); ttk.Spinbox(l_row, from_=0, to=59, textvariable=self.late_sec, width=5).pack(side='left')
        ttk.Label(l_row, text="sec").pack(side='left', padx=2)
        
        tk.Button(self.setup_frame, text="▶ START SESSION", font=('Arial', 11, 'bold'), bg='#4CAF50', fg='white', command=self.start_session).pack(pady=10)

        self.active_frame = ttk.LabelFrame(main_container, text="Session In Progress", padding=15)
        self.active_info_label = tk.Label(self.active_frame, text="", font=('Arial', 10)); self.active_info_label.pack()
        self.status_box = tk.Label(self.active_frame, text="ATTENDANCE OPEN", font=('Arial', 14, 'bold'), bg='#4CAF50', fg='white', width=40, pady=5); self.status_box.pack(pady=10)
        self.timer_label = tk.Label(self.active_frame, text="00:00", font=('Courier New', 24, 'bold'), fg='#333'); self.timer_label.pack()
        self.phase_label = tk.Label(self.active_frame, text="(Normal Phase)", fg='gray'); self.phase_label.pack()
        input_frame = ttk.Frame(self.active_frame); input_frame.pack(pady=15)
        ttk.Label(input_frame, text="SCAN RFID / KEY:", font=('Arial', 12, 'bold')).pack(side='left', padx=5)
        self.scan_entry = ttk.Entry(input_frame, font=('Arial', 14), width=30); self.scan_entry.pack(side='left', padx=5); self.scan_entry.bind('<Return>', lambda e: self.process_scan())
        tk.Button(self.active_frame, text="⏹ Close Session Early", bg='#f44336', fg='white', command=self.close_session).pack(pady=5)

        list_label = tk.Label(main_container, text="📋 Live Class List", font=('Arial', 12, 'bold')); list_label.pack(pady=(10, 5))
        cols = ('Key', 'Name', 'Role', 'Status', 'Time')
        self.live_tree = ttk.Treeview(main_container, columns=cols, show='headings', height=15)
        self.live_tree.heading('Key', text='Person Key'); self.live_tree.heading('Name', text='Name'); self.live_tree.heading('Role', text='Role'); self.live_tree.heading('Status', text='Status'); self.live_tree.heading('Time', text='Scan Time')
        self.live_tree.column('Key', width=100); self.live_tree.column('Name', width=200); self.live_tree.column('Role', width=120); self.live_tree.column('Status', width=100); self.live_tree.column('Time', width=150)
        self.live_tree.tag_configure('present', background='#C8E6C9'); self.live_tree.tag_configure('late', background='#FFE0B2'); self.live_tree.tag_configure('absent', background='white')
        scrollbar = ttk.Scrollbar(main_container, orient="vertical", command=self.live_tree.yview); self.live_tree.configure(yscrollcommand=scrollbar.set); self.live_tree.pack(side='left', fill='both', expand=True); scrollbar.pack(side='right', fill='y')
        self.active_frame.pack_forget() 

    def tick_clock(self):
        try:
            current_time = datetime.now().strftime("%H:%M:%S")
            self.rt_clock_label.config(text=current_time)
            self.root.after(1000, self.tick_clock)
        except: pass

    def build_history_tab(self):
        top_frame = ttk.Frame(self.history_frame); top_frame.pack(fill='x', padx=20, pady=10)
        tk.Label(top_frame, text="Attendance History", font=('Arial', 16, 'bold')).pack(side='left')
        btn_frame = ttk.Frame(top_frame); btn_frame.pack(side='right')
        tk.Button(btn_frame, text="📊 Excel", bg='#2196F3', fg='white', command=self.export_history).pack(side='left', padx=5)
        tk.Button(btn_frame, text="🗑️ Delete", bg='#f44336', fg='white', command=self.delete_history).pack(side='left', padx=5)
        tk.Button(btn_frame, text="🔄 Refresh", command=self.load_history).pack(side='left', padx=5)
        cols = ('ID', 'Date', 'Title', 'Starter', 'Total', 'Status')
        self.hist_tree = ttk.Treeview(self.history_frame, columns=cols, show='headings')
        for c in cols: self.hist_tree.heading(c, text=c)
        self.hist_tree.column('ID', width=50); self.hist_tree.pack(fill='both', expand=True, padx=20, pady=10); self.load_history()

    def start_session(self):
        starter = self.starter_key_entry.get().strip(); title = self.session_title_entry.get().strip()
        if not starter or not title: messagebox.showerror("Error", "Required fields empty."); return
        self.cursor.execute("SELECT name FROM people WHERE key=?", (starter,)); res = self.cursor.fetchone()
        if not res: messagebox.showerror("Error", "Starter not found."); return
        now = datetime.now()
        try:
            if self.time_mode.get() == "countdown": normal_dur = int(self.main_min.get())*60 + int(self.main_sec.get())
            else:
                end_t = now.replace(hour=int(self.end_hour.get()), minute=int(self.end_minute.get()), second=0, microsecond=0)
                if end_t <= now: end_t += timedelta(days=1)
                normal_dur = int((end_t - now).total_seconds())
            late_dur = int(self.late_min.get())*60 + int(self.late_sec.get())
            if normal_dur <= 0: messagebox.showerror("Error", "Invalid duration."); return
        except: messagebox.showerror("Error", "Invalid time."); return
        self.session_end_time = now + timedelta(seconds=normal_dur); self.late_end_time = self.session_end_time + timedelta(seconds=late_dur); self.current_phase = 'normal'
        self.cursor.execute('''INSERT INTO sessions (starter_key, title, description, date_str, start_time, normal_end_time, final_end_time, normal_duration_sec, late_duration_sec) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)''', (starter, title, self.session_desc_entry.get(), now.strftime("%Y-%m-%d"), now, self.session_end_time, self.late_end_time, normal_dur, late_dur))
        self.conn.commit(); self.active_session = self.cursor.lastrowid; self.setup_frame.pack_forget(); self.active_frame.pack(fill='x', pady=10); self.active_info_label.config(text=f"Session: {title}")
        self.scan_map.clear()
        for item in self.live_tree.get_children(): self.live_tree.delete(item)
        self.cursor.execute("SELECT key, name, personality FROM people ORDER BY name")
        for p in self.cursor.fetchall():
            iid = self.live_tree.insert('', 'end', values=(p[0], p[1], p[2], "Absent", ""), tags=('absent',)); self.scan_map[p[0]] = iid 
        self.scan_entry.focus(); self.update_timer()

    def update_timer(self):
        if not self.active_session: return
        now = datetime.now()
        if self.current_phase == 'normal':
            rem = (self.session_end_time - now).total_seconds()
            if rem <= 0: self.current_phase = 'late'; self.status_box.config(text="🟠 LATE CLOCK-IN", bg='#FF9800'); rem = (self.late_end_time - now).total_seconds()
            else: self.status_box.config(bg='#4CAF50')
        else:
            rem = (self.late_end_time - now).total_seconds()
            if rem <= 0: self.close_session(); return
        m, s = divmod(int(rem), 60); self.timer_label.config(text=f"{m:02d}:{s:02d}"); self.session_timer = self.root.after(1000, self.update_timer)

    def process_scan(self):
        if not self.active_session: return
        key = self.scan_entry.get().strip(); self.scan_entry.delete(0, 'end')
        if key not in self.scan_map: return 
        item_id = self.scan_map[key]; vals = self.live_tree.item(item_id)['values']
        if vals[3] != "Absent": return
        st = "Present" if self.current_phase == 'normal' else "Late"; tag = 'present' if self.current_phase == 'normal' else 'late'; now_t = datetime.now().strftime("%H:%M:%S")
        self.live_tree.item(item_id, values=(vals[0], vals[1], vals[2], st, now_t), tags=(tag,)); self.live_tree.see(item_id)
        self.cursor.execute('''INSERT INTO attendance (session_id, person_key, scan_time, status) VALUES (?, ?, ?, ?)''', (self.active_session, key, datetime.now(), st)); self.conn.commit()

    def close_session(self):
        if self.session_timer: self.root.after_cancel(self.session_timer)
        self.active_session = None; self.cursor.execute("UPDATE sessions SET status='closed' WHERE id=?", (self.active_session,))
        self.conn.commit(); self.active_frame.pack_forget(); self.setup_frame.pack(fill='x', pady=10); messagebox.showinfo("Finished", "Session Closed."); self.load_history()

    def load_history(self):
        for item in self.hist_tree.get_children(): self.hist_tree.delete(item)
        self.cursor.execute('SELECT s.id, s.date_str, s.title, p.name, (s.normal_duration_sec + s.late_duration_sec), s.status FROM sessions s JOIN people p ON s.starter_key = p.key ORDER BY s.id DESC')
        for r in self.cursor.fetchall(): self.hist_tree.insert('', 'end', values=r)

    def delete_history(self):
        sel = self.hist_tree.selection()
        if not sel or not messagebox.askyesno("Confirm", "Delete selected session(s)?"): return
        for i in sel:
            sid = self.hist_tree.item(i)['values'][0]
            self.cursor.execute("DELETE FROM attendance WHERE session_id=?", (sid,)); self.cursor.execute("DELETE FROM sessions WHERE id=?", (sid,))
        self.conn.commit(); self.load_history()

    def export_history(self):
        sel = self.hist_tree.selection()
        if not sel: return
        fn = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")])
        if not fn: return
        wb = Workbook(); wb.remove(wb.active)
        for i in sel:
            v = self.hist_tree.item(i)['values']; sid = v[0]
            ws = wb.create_sheet(title=f"{sid}-{v[2]}"[:30].replace(":",""))
            ws.append(["Report", v[2], v[1], v[3]]); ws.append(["Name", "Role", "Time", "Status", "Key"])
            self.cursor.execute('SELECT p.name, p.personality, a.scan_time, a.status, p.key FROM attendance a JOIN people p ON a.person_key = p.key WHERE a.session_id = ?', (sid,))
            for r in self.cursor.fetchall(): ws.append([r[0], r[1], r[2], r[3], r[4]])
        wb.save(fn); messagebox.showinfo("Success", "Exported!")

    def build_new_tab(self):
        tk.Label(self.new_frame, text="Add New Person", font=('Arial', 16, 'bold'), pady=10).pack()
        imp = ttk.LabelFrame(self.new_frame, text="Import", padding=10); imp.pack(pady=10, padx=40, fill='x')
        tk.Button(imp, text="📥 Import", bg='#2196F3', fg='white', command=self.import_from_excel).pack(side='left', padx=10)
        tk.Button(imp, text="📄 Template", bg='#9E9E9E', fg='white', command=self.download_template).pack(side='left', padx=10)
        f = ttk.LabelFrame(self.new_frame, text="Manual", padding=20); f.pack(pady=10, padx=40)
        ttk.Label(f, text="Key:").grid(row=0, column=0, sticky='w'); self.new_key_entry = ttk.Entry(f, width=40); self.new_key_entry.grid(row=0, column=1, padx=10, pady=5)
        ttk.Label(f, text="Name:").grid(row=1, column=0, sticky='w'); self.new_name_entry = ttk.Entry(f, width=40); self.new_name_entry.grid(row=1, column=1, padx=10, pady=5)
        ttk.Label(f, text="DOB:").grid(row=2, column=0, sticky='w'); self.new_dob_entry = ttk.Entry(f, width=15); self.new_dob_entry.grid(row=2, column=1, sticky='w', padx=10); self.new_dob_entry.bind('<KeyRelease>', self.calculate_age)
        self.age_display = tk.Label(f, text="--", font=('Arial', 11, 'bold'), fg='#2196F3'); self.age_display.grid(row=3, column=1, sticky='w', padx=10)
        ttk.Label(f, text="Role:").grid(row=4, column=0, sticky='w'); self.new_personality_entry = ttk.Entry(f, width=40); self.new_personality_entry.grid(row=4, column=1, padx=10, pady=5)
        bt = ttk.Frame(self.new_frame); bt.pack(pady=20)
        tk.Button(bt, text="Add", bg='#4CAF50', fg='white', width=15, command=self.add_person).pack(side='left', padx=5)
        tk.Button(bt, text="Clear", bg='#FF9800', fg='white', width=15, command=self.clear_new_form).pack(side='left', padx=5)

    def download_template(self):
        fn = filedialog.asksaveasfilename(defaultextension=".xlsx", initialfile="Template.xlsx")
        if not fn: return
        wb = Workbook(); ws = wb.active; ws.append(["Key", "Name", "DOB", "Role"]); wb.save(fn)

    def import_from_excel(self):
        fn = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx")])
        if not fn: return
        try:
            wb = load_workbook(fn); ws = wb.active
            for r in ws.iter_rows(min_row=2, values_only=True):
                if not r[0]: continue
                try:
                    dt = r[2] if isinstance(r[2], datetime) else datetime.strptime(str(r[2]), '%Y-%m-%d')
                    age = datetime.now().year - dt.year
                    self.cursor.execute('INSERT INTO people (key, name, dob, age, personality) VALUES (?, ?, ?, ?, ?)', (str(r[0]), str(r[1]), dt.strftime('%Y-%m-%d'), age, r[3]))
                except: continue
            self.conn.commit(); self.load_people_list(); messagebox.showinfo("Done", "Imported.")
        except: messagebox.showerror("Error", "Failed.")

    def calculate_age(self, event=None):
        try:
            dt = datetime.strptime(self.new_dob_entry.get(), '%Y-%m-%d')
            self.age_display.config(text=f"{datetime.now().year - dt.year} years old")
        except: self.age_display.config(text="--")

    def add_person(self):
        k, n, d, r = self.new_key_entry.get(), self.new_name_entry.get(), self.new_dob_entry.get(), self.new_personality_entry.get()
        if not k or not n or not d: return
        try:
            age = datetime.now().year - datetime.strptime(d, '%Y-%m-%d').year
            self.cursor.execute('INSERT INTO people (key, name, dob, age, personality) VALUES (?, ?, ?, ?, ?)', (k, n, d, age, r))
            self.conn.commit(); self.clear_new_form(); self.load_people_list(); messagebox.showinfo("Success", "Added.")
        except: messagebox.showerror("Error", "Failed.")

    def clear_new_form(self):
        for e in [self.new_key_entry, self.new_name_entry, self.new_dob_entry, self.new_personality_entry]: e.delete(0, 'end')
        self.age_display.config(text="--")

    # --- LIST / VIEW PEOPLE (UPDATED WITH SORT & FILTER) ---

    def build_list_tab(self):
        top_frame = ttk.Frame(self.list_frame)
        top_frame.pack(fill='x', padx=20, pady=10)
        tk.Label(top_frame, text="Registered People", font=('Arial', 16, 'bold')).pack(side='left')
        
        # Filter Search Bar
        filter_frame = ttk.Frame(top_frame)
        filter_frame.pack(side='left', padx=30)
        ttk.Label(filter_frame, text="🔍 Filter:").pack(side='left', padx=5)
        self.filter_var = tk.StringVar()
        self.filter_var.trace_add("write", lambda *args: self.load_people_list())
        self.filter_entry = ttk.Entry(filter_frame, textvariable=self.filter_var, width=30)
        self.filter_entry.pack(side='left')

        # Action Buttons
        btn_frame = ttk.Frame(top_frame); btn_frame.pack(side='right')
        tk.Button(btn_frame, text="✏️ Edit", bg='#2196F3', fg='white', command=self.edit_selected_person).pack(side='left', padx=5)
        tk.Button(btn_frame, text="🗑️ Delete", bg='#f44336', fg='white', command=self.delete_selected_person).pack(side='left', padx=5)
        tk.Button(btn_frame, text="🔄 Refresh", command=self.load_people_list).pack(side='left', padx=5)

        # Treeview with Sorting logic
        cols = ('Key', 'Name', 'DOB', 'Age', 'Role')
        self.pl_tree = ttk.Treeview(self.list_frame, columns=cols, show='headings')
        
        # Bind sorting to headers (except Key)
        for col in cols:
            if col == 'Key':
                self.pl_tree.heading(col, text=col)
            else:
                # We use a lambda to pass the specific column name to the sort function
                self.pl_tree.heading(col, text=f"{col} ↕", command=lambda _c=col: self.sort_by_column(_c))
            
        self.pl_tree.column('Key', width=100); self.pl_tree.column('Name', width=200); self.pl_tree.column('DOB', width=100); self.pl_tree.column('Age', width=50); self.pl_tree.column('Role', width=150)
        self.pl_tree.pack(fill='both', expand=True, padx=20, pady=20)
        
        self.pl_tree.bind("<Double-1>", lambda e: self.edit_selected_person())
        self.pl_tree.bind("<Button-3>", self.show_context_menu)
        self.list_menu = Menu(self.root, tearoff=0)
        self.list_menu.add_command(label="✏️ Edit", command=self.edit_selected_person)
        self.list_menu.add_command(label="🗑️ Delete", command=self.delete_selected_person)
        self.load_people_list()

    def sort_by_column(self, col):
        """Toggle sorting order for a specific column and refresh list"""
        # Toggle current sort state
        self.sort_state[col] = not self.sort_state[col]
        # Reload list (the SQL query handles the logic)
        self.load_people_list(sort_col=col)

    def load_people_list(self, sort_col=None):
        """Fetches people from DB applying filters and sorting"""
        for i in self.pl_tree.get_children(): self.pl_tree.delete(i)
        
        filter_text = f"%{self.filter_var.get()}%"
        
        # Base Query with Filter logic
        query = '''
            SELECT key, name, dob, age, personality 
            FROM people 
            WHERE (name LIKE ? OR dob LIKE ? OR age LIKE ? OR personality LIKE ?)
        '''
        
        # Append Sorting logic
        if sort_col:
            direction = "ASC" if self.sort_state[sort_col] else "DESC"
            # Map friendly column name to DB column name
            db_col = {"Name": "name", "DOB": "dob", "Age": "age", "Role": "personality"}[sort_col]
            query += f" ORDER BY {db_col} {direction}"
        else:
            query += " ORDER BY name ASC"

        self.cursor.execute(query, (filter_text, filter_text, filter_text, filter_text))
        for r in self.cursor.fetchall(): self.pl_tree.insert('', 'end', iid=r[0], values=r)

    def show_context_menu(self, event):
        item = self.pl_tree.identify_row(event.y)
        if item: self.pl_tree.selection_set(item); self.list_menu.tk_popup(event.x_root, event.y_root)

    def delete_selected_person(self):
        sel = self.pl_tree.selection()
        if not sel: return
        key = sel[0]; name = self.pl_tree.item(key)['values'][1]
        if not messagebox.askyesno("Confirm", f"Delete {name}?"): return
        try:
            self.cursor.execute("DELETE FROM attendance WHERE person_key=?", (key,))
            self.cursor.execute("DELETE FROM people WHERE key=?", (key,)); self.conn.commit()
            messagebox.showinfo("Success", "Deleted."); self.load_people_list()
        except: messagebox.showerror("Error", "Failed.")

    def edit_selected_person(self):
        sel = self.pl_tree.selection()
        if not sel: return
        key = sel[0]; vals = self.pl_tree.item(key)['values']
        ed = tk.Toplevel(self.root); ed.title("Edit"); ed.geometry("400x350")
        tk.Label(ed, text="Edit Details", font=("Arial", 14, "bold")).pack(pady=10)
        f = tk.Frame(ed); f.pack(pady=10, padx=20)
        tk.Label(f, text="Name:").grid(row=1, column=0); n = tk.Entry(f); n.insert(0, vals[1]); n.grid(row=1, column=1, pady=5)
        tk.Label(f, text="DOB:").grid(row=2, column=0); d = tk.Entry(f); d.insert(0, vals[2]); d.grid(row=2, column=1, pady=5)
        tk.Label(f, text="Role:").grid(row=3, column=0); r = tk.Entry(f); r.insert(0, vals[4]); r.grid(row=3, column=1, pady=5)
        def save():
            try:
                dt = datetime.strptime(d.get(), '%Y-%m-%d'); age = datetime.now().year - dt.year
                self.cursor.execute('UPDATE people SET name=?, dob=?, age=?, personality=? WHERE key=?', (n.get(), d.get(), age, r.get(), key))
                self.conn.commit(); self.load_people_list(); ed.destroy(); messagebox.showinfo("Success", "Updated.")
            except: messagebox.showerror("Error", "Failed.")
        tk.Button(ed, text="Save", bg="#4CAF50", fg="white", command=save).pack(pady=20)

if __name__ == "__main__":
    root = tk.Tk()
    app = AttendanceApp(root)
    root.mainloop()